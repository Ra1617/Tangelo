"""
Excel Agent — Creates and edits Excel spreadsheets (.xlsx).
Uses openpyxl for creation and formatting, supports charts.
"""

import os
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

from agents.base_agent import BaseAgent, AgentResult
from config import OUTPUT_DIR

logger = logging.getLogger("excel_agent")


class ExcelAgent(BaseAgent):
    name = "excel"
    description = "Create Excel spreadsheets, add data, charts, and formatting"
    capabilities = ["create_spreadsheet", "add_data", "add_chart", "apply_formatting"]

    def execute(self, action: str, args: dict) -> AgentResult:
        self._validate_action(action)

        if action == "create_spreadsheet":
            return self._create_spreadsheet(**args)
        elif action == "add_data":
            return self._add_data(**args)
        elif action == "add_chart":
            return self._add_chart(**args)
        elif action == "apply_formatting":
            return self._apply_formatting(**args)

        return AgentResult(success=False, message=f"Unknown action: {action}")

    # ── Actions ──────────────────────────────

    def _create_spreadsheet(self, filename: str, sheet_name: str,
                            headers: list, rows: list) -> AgentResult:
        """Create a new .xlsx with headers and data rows."""
        try:
            if not filename:
                import time
                filename = f"spreadsheet_{int(time.time())}.xlsx"
            if not sheet_name:
                sheet_name = "Sheet1"
            if not headers:
                headers = []
            if not rows:
                rows = []

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Header style
            header_fill = PatternFill("solid", fgColor="1A1A2E")
            header_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
            thin_border = Border(
                bottom=Side(style="thin", color="4F8EF7")
            )

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(str(header)) + 4)

            # Data rows with alternating colors
            data_font = Font(size=10, name="Calibri")
            alt_fill = PatternFill("solid", fgColor="F0F4FF")

            for row_idx, row in enumerate(rows, start=2):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal="left")
                    if row_idx % 2 == 0:
                        cell.fill = alt_fill

            # Auto-filter
            if headers:
                last_col = get_column_letter(len(headers))
                ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

            # Freeze header row
            ws.freeze_panes = "A2"

            # Save
            if not filename.endswith(".xlsx"):
                filename += ".xlsx"
            path = os.path.join(OUTPUT_DIR, filename)
            wb.save(path)

            logger.info("Created spreadsheet: %s", path)
            return AgentResult(
                success=True,
                message=f"✅ Spreadsheet created: {filename} ({len(rows)} rows)",
                output_file=path
            )

        except Exception as e:
            logger.exception("Failed to create spreadsheet")
            return AgentResult(success=False, message=f"❌ Excel error: {e}")

    def _add_data(self, filename: str, sheet_name: str, rows: list) -> AgentResult:
        """Append data rows to an existing spreadsheet."""
        try:
            if not filename.endswith(".xlsx"):
                filename += ".xlsx"
            path = os.path.join(OUTPUT_DIR, filename)

            if not os.path.exists(path):
                return AgentResult(success=False, message=f"❌ File not found: {filename}")

            wb = openpyxl.load_workbook(path)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)

            for row in rows:
                ws.append(row)

            wb.save(path)
            logger.info("Added %d rows to %s", len(rows), path)
            return AgentResult(
                success=True,
                message=f"✅ Added {len(rows)} rows to {filename}",
                output_file=path
            )

        except Exception as e:
            logger.exception("Failed to add data")
            return AgentResult(success=False, message=f"❌ Add data error: {e}")

    def _add_chart(self, filename: str, sheet_name: str,
                   chart_type: str = "bar", data_range: str = "") -> AgentResult:
        """Add a chart to the spreadsheet."""
        try:
            if not filename.endswith(".xlsx"):
                filename += ".xlsx"
            path = os.path.join(OUTPUT_DIR, filename)

            if not os.path.exists(path):
                return AgentResult(success=False, message=f"❌ File not found: {filename}")

            wb = openpyxl.load_workbook(path)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

            # Determine data range automatically if not provided
            max_row = ws.max_row
            max_col = ws.max_column

            if max_row < 2 or max_col < 2:
                return AgentResult(success=False, message="❌ Not enough data for a chart")

            # Create chart based on type
            chart_map = {"bar": BarChart, "line": LineChart, "pie": PieChart}
            chart_cls = chart_map.get(chart_type.lower(), BarChart)
            chart = chart_cls()

            chart.title = f"{sheet_name} Chart"
            chart.style = 10
            chart.width = 20
            chart.height = 12

            # Data reference (columns 2+ as data, column 1 as categories)
            data_ref = Reference(ws, min_col=2, min_row=1, max_col=max_col, max_row=max_row)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=max_row)

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)

            # Place chart below data
            chart_cell = f"A{max_row + 3}"
            ws.add_chart(chart, chart_cell)

            wb.save(path)
            logger.info("Added %s chart to %s", chart_type, path)
            return AgentResult(
                success=True,
                message=f"✅ {chart_type.capitalize()} chart added to {filename}",
                output_file=path
            )

        except Exception as e:
            logger.exception("Failed to add chart")
            return AgentResult(success=False, message=f"❌ Chart error: {e}")

    def _apply_formatting(self, filename: str, sheet_name: str,
                          style: str = "professional") -> AgentResult:
        """Apply visual formatting styles to the spreadsheet."""
        try:
            if not filename.endswith(".xlsx"):
                filename += ".xlsx"
            path = os.path.join(OUTPUT_DIR, filename)

            if not os.path.exists(path):
                return AgentResult(success=False, message=f"❌ File not found: {filename}")

            wb = openpyxl.load_workbook(path)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

            # Apply professional styling
            header_fill = PatternFill("solid", fgColor="2C3E50")
            header_font = Font(color="FFFFFF", bold=True, size=12, name="Calibri")
            accent_border = Border(
                bottom=Side(style="medium", color="3498DB"),
                top=Side(style="thin", color="BDC3C7"),
            )

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = accent_border
                cell.alignment = Alignment(horizontal="center")

            wb.save(path)
            logger.info("Applied formatting to %s", path)
            return AgentResult(
                success=True,
                message=f"✅ Formatting applied to {filename}",
                output_file=path
            )

        except Exception as e:
            logger.exception("Failed to apply formatting")
            return AgentResult(success=False, message=f"❌ Formatting error: {e}")

"""
AI Office Assistant — Premium Desktop GUI
Warm, modern AI-tool theme with orange, amber, and soft cream tones.
Card-based layout with soft shadows. Communicates with FastAPI backend.
"""

import sys
import os
import threading
import time
import subprocess

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog as fd

import requests

from config import API_HOST, API_PORT

# ─────────────────────────────────────────────
#  THEME — Warm Orange / Cream
# ─────────────────────────────────────────────

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# ── Color Palette ──
PRIMARY        = "#E85D0F"        # strong orange — header, primary buttons
SECONDARY      = "#F29C52"        # light orange — borders, hover, accents
BG_MAIN        = "#F7EADB"        # light cream — main background
BG_PANEL       = "#F3E0CD"        # slightly darker cream — cards, panels
ACCENT_PINK    = "#E85D75"        # soft pink/red — notifications, active
TEXT_PRIMARY    = "#2B1B14"        # dark brown/black
TEXT_SECONDARY  = "#6B4A3A"        # muted brown
BORDER         = "#E0A96D"        # soft orange border
SUCCESS        = "#00C853"        # green for online status
ERROR_RED      = "#D32F2F"        # red for errors
WHITE          = "#FFFFFF"
INPUT_BG       = "#FFFFFF"        # white input boxes

FONT_FAMILY    = "Segoe UI"
MONO_FAMILY    = "Consolas"

API_BASE = f"http://{API_HOST}:{API_PORT}"


# ─────────────────────────────────────────────
#  CHAT BUBBLE WIDGET
# ─────────────────────────────────────────────

class ChatBubble(ctk.CTkFrame):
    """A warm-styled chat message bubble."""

    def __init__(self, parent, text: str, role: str,
                 file_paths: list | None = None, **kwargs):
        if role == "user":
            bg = "#FCEADE"       # warm peach for user
            border_clr = PRIMARY
        else:
            bg = WHITE           # white for agent
            border_clr = BORDER

        super().__init__(
            parent, fg_color=bg, corner_radius=12,
            border_width=2, border_color=border_clr,
            **kwargs
        )

        # Role label
        label_text = "  You" if role == "user" else "  🤖 Agent"
        label_color = PRIMARY if role == "user" else "#C0392B"
        ctk.CTkLabel(
            self, text=label_text,
            font=ctk.CTkFont(family=FONT_FAMILY, size=14, weight="bold"),
            text_color=label_color,
        ).pack(anchor="w", padx=16, pady=(12, 4))

        # Message text
        ctk.CTkLabel(
            self, text=text,
            font=ctk.CTkFont(family=FONT_FAMILY, size=15),
            text_color=TEXT_PRIMARY,
            wraplength=520,
            justify="left",
            anchor="w",
        ).pack(anchor="w", padx=16, pady=(0, 6))

        # File open buttons
        if file_paths:
            for fp in file_paths:
                if os.path.exists(fp):
                    btn = ctk.CTkButton(
                        self,
                        text=f"  📂  {os.path.basename(fp)}",
                        font=ctk.CTkFont(family=MONO_FAMILY, size=13),
                        fg_color=SECONDARY,
                        hover_color=PRIMARY,
                        text_color=WHITE,
                        height=32, corner_radius=8,
                        anchor="w",
                        command=lambda p=fp: os.startfile(p) if sys.platform == "win32"
                            else subprocess.call(["xdg-open", p])
                    )
                    btn.pack(anchor="w", padx=16, pady=3)

        # Bottom padding
        ctk.CTkLabel(self, text="", height=6).pack()


# ─────────────────────────────────────────────
#  STATUS LOG WIDGET (Bottom-right card)
# ─────────────────────────────────────────────

class StatusLog(ctk.CTkFrame):
    """Live execution log — shows think/act/observe/update phases."""

    def __init__(self, parent, **kwargs):
        super().__init__(
            parent, fg_color=BG_PANEL, corner_radius=12,
            border_width=2, border_color=BORDER, **kwargs
        )

        # Header row
        header = ctk.CTkFrame(self, fg_color="transparent", height=40)
        header.pack(fill="x", padx=16, pady=(14, 0))

        ctk.CTkLabel(
            header, text="⚡ Execution Log",
            font=ctk.CTkFont(family=FONT_FAMILY, size=16, weight="bold"),
            text_color=PRIMARY,
        ).pack(side="left")

        ctk.CTkButton(
            header, text="Clear", width=60, height=28,
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            fg_color="transparent", text_color=TEXT_SECONDARY,
            hover_color=SECONDARY, corner_radius=8,
            command=self.clear,
        ).pack(side="right")

        # Scrollable log area
        self._log_text = ctk.CTkTextbox(
            self, height=160,
            font=ctk.CTkFont(family=MONO_FAMILY, size=13),
            fg_color=WHITE,
            text_color=TEXT_SECONDARY,
            border_width=1,
            border_color=BORDER,
            corner_radius=10,
            wrap="word",
            state="disabled",
        )
        self._log_text.pack(fill="both", expand=True, padx=16, pady=(10, 16))

    def log(self, message: str, color: str | None = None):
        self._log_text.configure(state="normal")
        self._log_text.insert("end", f"{message}\n")
        self._log_text.configure(state="disabled")
        self._log_text.see("end")

    def clear(self):
        self._log_text.configure(state="normal")
        self._log_text.delete("1.0", "end")
        self._log_text.configure(state="disabled")


# ─────────────────────────────────────────────
#  PLAN VIEWER WIDGET (Top-right card)
# ─────────────────────────────────────────────

class PlanViewer(ctk.CTkFrame):
    """Shows the current JSON execution plan with step statuses."""

    def __init__(self, parent, **kwargs):
        super().__init__(
            parent, fg_color=BG_PANEL, corner_radius=12,
            border_width=2, border_color=BORDER, **kwargs
        )

        ctk.CTkLabel(
            self, text="📋 Execution Plan",
            font=ctk.CTkFont(family=FONT_FAMILY, size=16, weight="bold"),
            text_color=PRIMARY,
        ).pack(anchor="w", padx=16, pady=(14, 6))

        self._steps_frame = ctk.CTkFrame(self, fg_color="transparent")
        self._steps_frame.pack(fill="x", padx=16, pady=(0, 14))

        self._show_empty()

    def _show_empty(self):
        ctk.CTkLabel(
            self._steps_frame,
            text="No active plan",
            font=ctk.CTkFont(family=FONT_FAMILY, size=14),
            text_color=TEXT_SECONDARY,
        ).pack(anchor="w", pady=4)

    def update_plan(self, plan_data: dict | None):
        for w in self._steps_frame.winfo_children():
            w.destroy()

        if not plan_data or "steps" not in plan_data:
            self._show_empty()
            return

        # Goal line
        goal = plan_data.get("goal", "")
        if goal:
            ctk.CTkLabel(
                self._steps_frame,
                text=f"🎯 {goal}",
                font=ctk.CTkFont(family=FONT_FAMILY, size=14),
                text_color=TEXT_PRIMARY,
                wraplength=300,
            ).pack(anchor="w", pady=(0, 8))

        # Steps
        status_icons = {
            "pending": "⏳", "running": "🔄",
            "success": "✅", "failed": "❌", "skipped": "⏭",
        }
        for step in plan_data.get("steps", []):
            icon = status_icons.get(step.get("status", "pending"), "⏳")
            text = f"{icon}  Step {step['id']}: {step['tool']}.{step['action']}"
            if step.get("status") == "success":
                color = SUCCESS
            elif step.get("status") == "failed":
                color = ERROR_RED
            else:
                color = TEXT_SECONDARY

            ctk.CTkLabel(
                self._steps_frame,
                text=text,
                font=ctk.CTkFont(family=MONO_FAMILY, size=13),
                text_color=color,
                anchor="w",
            ).pack(anchor="w", pady=2)


# ─────────────────────────────────────────────
#  MAIN APPLICATION
# ─────────────────────────────────────────────

class AgentOSApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("AI Office Assistant — Powered by Ollama")
        self.geometry("1200x820")
        self.configure(fg_color=BG_MAIN)
        self.minsize(950, 680)

        self.attached_file = None
        self._build_ui()
        self._check_health()

    # ══════════════════════════════════════════
    #  BUILD UI
    # ══════════════════════════════════════════

    def _build_ui(self):

        # ═══════════════════════════════════════
        #  HEADER BAR — solid orange
        # ═══════════════════════════════════════
        header = ctk.CTkFrame(self, fg_color=PRIMARY, corner_radius=0, height=64)
        header.pack(fill="x", side="top")
        header.pack_propagate(False)

        # Title (left)
        ctk.CTkLabel(
            header,
            text="🟠  AI Office Assistant  v1.0",
            font=ctk.CTkFont(family=FONT_FAMILY, size=22, weight="bold"),
            text_color=WHITE,
        ).pack(side="left", padx=24, pady=14)

        # Status indicator (right-most)
        self._status_indicator = ctk.CTkLabel(
            header,
            text="● checking...",
            font=ctk.CTkFont(family=MONO_FAMILY, size=13, weight="bold"),
            text_color="#FFD9B3",
        )
        self._status_indicator.pack(side="right", padx=24, pady=14)

        # Reset button (right)
        ctk.CTkButton(
            header,
            text="↺  Reset",
            font=ctk.CTkFont(family=FONT_FAMILY, size=14, weight="bold"),
            fg_color=WHITE,
            hover_color="#FFD9B3",
            text_color=PRIMARY,
            border_width=0,
            width=90, height=34,
            corner_radius=10,
            command=self._reset_chat,
        ).pack(side="right", padx=6, pady=14)

        # ═══════════════════════════════════════
        #  INPUT BAR — bottom (build before body so it sits at bottom)
        # ═══════════════════════════════════════
        input_bar = ctk.CTkFrame(
            self, fg_color=BG_MAIN, corner_radius=0, height=90,
        )
        input_bar.pack(fill="x", side="bottom")
        input_bar.pack_propagate(False)

        # Attachment label
        self._attach_label = ctk.CTkLabel(
            input_bar, text="",
            font=ctk.CTkFont(family=FONT_FAMILY, size=12),
            text_color=PRIMARY,
        )
        self._attach_label.pack(side="top", anchor="w", padx=24, pady=(6, 0))

        input_inner = ctk.CTkFrame(input_bar, fg_color="transparent")
        input_inner.pack(fill="x", padx=24, pady=(4, 18))

        # Attach button
        ctk.CTkButton(
            input_inner,
            text="📎",
            font=ctk.CTkFont(size=20),
            fg_color=SECONDARY,
            hover_color=PRIMARY,
            text_color=WHITE,
            width=50, height=50,
            corner_radius=12,
            command=self._attach_file,
        ).pack(side="left", padx=(0, 10))

        # Text input
        self.input_box = ctk.CTkEntry(
            input_inner,
            height=50,
            font=ctk.CTkFont(family=FONT_FAMILY, size=15),
            fg_color=INPUT_BG,
            text_color="#000000",
            border_color=BORDER,
            border_width=2,
            corner_radius=12,
        )
        self.input_box.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.input_box.bind("<Return>", lambda e: self._send())

        # Send button
        self.send_btn = ctk.CTkButton(
            input_inner,
            text="Send  ▶",
            font=ctk.CTkFont(family=FONT_FAMILY, size=16, weight="bold"),
            fg_color=PRIMARY,
            hover_color=SECONDARY,
            text_color=WHITE,
            corner_radius=12,
            width=120, height=50,
            command=self._send,
        )
        self.send_btn.pack(side="right")

        # ═══════════════════════════════════════
        #  MAIN CONTENT AREA — two-panel layout
        # ═══════════════════════════════════════
        body = ctk.CTkFrame(self, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=16, pady=(12, 8))

        # ── LEFT PANEL: Agent / Chat ─────────
        left_panel = ctk.CTkFrame(
            body, fg_color=BG_PANEL, corner_radius=12,
            border_width=2, border_color=BORDER,
        )
        left_panel.pack(side="left", fill="both", expand=True, padx=(0, 8))

        # Agent panel header
        agent_header = ctk.CTkFrame(left_panel, fg_color="transparent", height=44)
        agent_header.pack(fill="x", padx=16, pady=(14, 0))
        agent_header.pack_propagate(False)

        ctk.CTkLabel(
            agent_header,
            text="🤖 Agent",
            font=ctk.CTkFont(family=FONT_FAMILY, size=18, weight="bold"),
            text_color=PRIMARY,
        ).pack(side="left")

        # Divider line
        ctk.CTkFrame(
            left_panel, fg_color=BORDER, height=2, corner_radius=1,
        ).pack(fill="x", padx=16, pady=(6, 0))

        # Chat scroll area
        self.chat_frame = ctk.CTkScrollableFrame(
            left_panel,
            fg_color=BG_PANEL,
            scrollbar_button_color=BORDER,
            scrollbar_button_hover_color=PRIMARY,
        )
        self.chat_frame.pack(fill="both", expand=True, padx=8, pady=(8, 12))

        # Welcome message
        self._add_bubble(
            "Hello! I'm your AI Office Assistant running 100% locally.\n\n"
            "I can help you:\n"
            "• 📝 Create Word documents & export to PDF\n"
            "• 📊 Build Excel spreadsheets with charts\n"
            "• 📧 Send emails via Outlook\n"
            "• 💻 Generate & run code files\n\n"
            "I use an agentic loop:\n"
            "Think → Act → Observe → Update\n\n"
            "Try: \"Create a Q1 sales spreadsheet and summary doc\"",
            role="agent"
        )

        # ── VERTICAL DIVIDER ─────────────────
        ctk.CTkFrame(
            body, fg_color=BORDER, width=2, corner_radius=1,
        ).pack(side="left", fill="y", padx=0, pady=10)

        # ── RIGHT PANEL: Plan + Log ──────────
        right_panel = ctk.CTkFrame(body, fg_color="transparent", width=370)
        right_panel.pack(side="right", fill="y", padx=(8, 0))
        right_panel.pack_propagate(False)

        # Execution Plan card (top)
        self.plan_viewer = PlanViewer(right_panel)
        self.plan_viewer.pack(fill="x", pady=(0, 10))

        # Execution Log card (bottom, expands)
        self.status_log = StatusLog(right_panel)
        self.status_log.pack(fill="both", expand=True)

    # ══════════════════════════════════════════
    #  HEALTH CHECK
    # ══════════════════════════════════════════

    def _check_health(self):
        def check():
            try:
                resp = requests.get(f"{API_BASE}/health", timeout=5)
                data = resp.json()
                ollama_ok = data.get("ollama_connected", False)
                model = data.get("model", "unknown")
                if ollama_ok:
                    self.after(0, lambda: self._status_indicator.configure(
                        text=f"● {model} · online",
                        text_color=SUCCESS,
                    ))
                else:
                    self.after(0, lambda: self._status_indicator.configure(
                        text="● Ollama offline",
                        text_color=ERROR_RED,
                    ))
            except Exception:
                self.after(0, lambda: self._status_indicator.configure(
                    text="● API starting...",
                    text_color="#FFD9B3",
                ))
                time.sleep(3)
                self.after(0, self._check_health)

        threading.Thread(target=check, daemon=True).start()

    # ══════════════════════════════════════════
    #  ACTIONS
    # ══════════════════════════════════════════

    def _on_enter(self, event):
        if not event.state & 0x1:
            self._send()
            return "break"

    def _send(self):
        text = self.input_box.get().strip()
        if not text:
            return

        self.input_box.delete(0, "end")

        display = text
        if self.attached_file:
            display = f"📎 [{os.path.basename(self.attached_file)}] {text}"
        self._add_bubble(display, role="user")

        self.send_btn.configure(state="disabled", text="  ⏳  ")
        self.status_log.log("━" * 38)
        self.status_log.log(f"📝 New request: {text[:80]}...")

        self.plan_viewer.update_plan(None)

        current_attached = self.attached_file

        threading.Thread(
            target=self._run_request,
            args=(text, current_attached),
            daemon=True,
        ).start()

        if self.attached_file:
            self.attached_file = None
            self._attach_label.configure(text="")

    def _run_request(self, prompt: str, attached_file: str | None = None):
        try:
            req_prompt = prompt
            if attached_file:
                content_preview = ""
                file_ext = os.path.splitext(attached_file)[1].lower()
                try:
                    if file_ext == ".docx":
                        import docx
                        doc = docx.Document(attached_file)
                        content_preview = "\n".join([p.text for p in doc.paragraphs])
                    elif file_ext in [".txt", ".csv", ".py", ".md", ".json"]:
                        with open(attached_file, "r", encoding="utf-8") as f:
                            content_preview = f.read()
                    elif file_ext == ".xlsx":
                        from openpyxl import load_workbook
                        wb = load_workbook(attached_file, data_only=True)
                        lines = []
                        for sheet in wb.worksheets:
                            lines.append(f"Sheet: {sheet.title}")
                            for row in sheet.iter_rows(values_only=True):
                                lines.append("\t".join(str(cell) for cell in row if cell is not None))
                        content_preview = "\n".join(lines)
                    else:
                        content_preview = f"(Cannot read contents for {file_ext} files. Path: {attached_file})"
                except Exception as e:
                    content_preview = f"(Error reading file content: {e})"
                
                if len(content_preview) > 4000:
                    content_preview = content_preview[:4000] + "\n...(truncated)"
                
                req_prompt = f"Attached document path: {os.path.abspath(attached_file)}\nAttached document content:\n{content_preview}\n\nTask: {req_prompt}"
            resp = requests.post(
                f"{API_BASE}/execute_sync",
                json={"prompt": req_prompt},
                timeout=300,
            )
            data = resp.json()

            for step_entry in data.get("steps_log", []):
                self.after(0, lambda m=step_entry["message"]: self.status_log.log(m))

            plan_data = data.get("plan")
            self.after(0, lambda: self.plan_viewer.update_plan(plan_data))

            message = data.get("message", "No response")
            files = data.get("files_created", [])
            self.after(0, lambda: self._add_bubble(
                message, role="agent", file_paths=files
            ))

        except requests.exceptions.ConnectionError:
            self.after(0, lambda: self._add_bubble(
                "❌ Cannot connect to API server. Try again in a moment.",
                role="agent"
            ))
        except Exception as e:
            self.after(0, lambda: self._add_bubble(
                f"❌ Error: {e}", role="agent"
            ))
        finally:
            self.after(0, lambda: self.send_btn.configure(
                state="normal", text="Send  ▶"
            ))

    def _add_bubble(self, text: str, role: str, file_paths: list | None = None):
        bubble = ChatBubble(
            self.chat_frame,
            text=text,
            role=role,
            file_paths=file_paths,
        )
        bubble.pack(fill="x", padx=8, pady=6, anchor="w")
        self.chat_frame.after(
            100, lambda: self.chat_frame._parent_canvas.yview_moveto(1.0)
        )

    def _reset_chat(self):
        for widget in self.chat_frame.winfo_children():
            widget.destroy()
        self.plan_viewer.update_plan(None)
        self.status_log.clear()
        self._add_bubble(
            "🔄 Session reset. How can I help you?",
            role="agent",
        )
        try:
            requests.post(
                f"{API_BASE}/execute_sync",
                json={"prompt": "__reset__"}, timeout=5
            )
        except Exception:
            pass

    def _attach_file(self):
        filetypes = [
            ("Common files", "*.pdf;*.docx;*.xlsx;*.pptx;*.txt;*.csv"),
            ("Images", "*.png;*.jpg;*.jpeg;*.gif;*.bmp"),
            ("All files", "*.*"),
        ]
        path = fd.askopenfilename(
            title="Attach a file",
            filetypes=filetypes,
        )
        if path:
            self.attached_file = path
            self._attach_label.configure(
                text=f"📎 {os.path.basename(path)}"
            )


# ─────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────

if __name__ == "__main__":
    app = AgentOSApp()
    app.mainloop()
